##########################################################
# Program Name: Project 2
#
# Author(s): Michael Cummings
#
# Date: 04/1/2025
#
# Description:
#A system that processes election data to calculate votes,
# percentages, and electoral results for Buggs Bunny (BB)
# and Road Runner (RR). It provides summaries, visualizations,
# and allows filtering of results by reporting deadlines.
###########################################################

#imported libraries
import csv
from csv import DictReader
from idlelib.editor import keynames
from openpyxl import workbook
from openpyxl.workbook import Workbook

from datetime import datetime
from scipy.constants import value

import matplotlib.pyplot as plt

#Initalize dictonaries
county_state_dict = {}
county_id_key_states_dict = {}
county_BBvotes_dict = {}
county_RRvotes_dict = {}
county_votes_dict = {}
states_votes = {}
states_BB_votes = {}
states_RR_votes = {}
states_votes_win_perc = {}
state_stateid = {}
state_num_counties = {}
county_return_time = {}
county_return_date = {}
county_votes_not_key = {}
county_lat_long_key = {}
county_BBvotes_not_key = {}
county_RRvotes_not_key = {}
state_electoral_votes = {}
state_canidate_win_perc_key = {}
state_canidate_elec_wins = {}
county_return_date_correct_format = {}
county_return_time_correct_format = {}
state_Buggs_votes = {}
state_Runner_votes = {}
state_list_county = {}
state_valid_county_votes = {}
state_perc_votes_placed = {}
state_num_reg_voters = {}

#initiliaze Lists
states = []
counties = []
counties_id = []
electoral_votes = []
valid_counties = []
valid_electoral_states = []
BB_electoral_votes_v2 = []
RR_electoral_votes_v2 = []


##########################################################
# Function/Method Name: read_county_data
#
# Author: Boris Pashkov
#
# Parameters: file_path (str) - The path to the county-level CSV file.
#
# Return Value: county_data (list) - A list of dictionaries with county-level
# vote data.
#
# Description:
# Reads and processes county-level data from a CSV file, parsing vote counts
# for each candidate (BB and RR) and other related data for each county.
##########################################################
def read_county_data(file_path):
    #open file and initilize reader
    with open(file_path, 'r', encoding='utf-8-sig') as file:
        reader = DictReader(file)
        #parse data and define varaibles
        for row in reader:
            county = row['county'].strip()
            county_id = row['county_fips']
            state = row['state_name']
            votes = row['Votes']
            RR_votes = row['RR Votes']
            BB_votes = row['BB Votes']
            ret_time = row['Return Time']
            ret_date = row['Return Date']
            lat = row['lat']
            lon = row['lng']
            #define key structure
            county_id_key = (county, county_id)
            #create county to votes dict without key format
            county_votes_not_key[county] = votes
            ##define key structure
            lat_lon_key = (lat, lon)
            #if statement to filter empty county or lat lon inputs
            if county and lat_lon_key:
                #create county to lat lon key dict
                county_lat_long_key[county] = lat_lon_key
            #crete county to return date and county to return time dicts
            county_return_time[county] = ret_time
            county_return_date[county] = ret_date
            #create county to state dict
            county_state_dict[county] = state
            #create county to BB_votes and county to RR_votes dict
            county_BBvotes_not_key[county] = BB_votes
            county_RRvotes_not_key[county] = RR_votes

            if county and county_id and county_id_key not in counties and county_id not in counties_id:
                counties_id.append(county_id)
                counties.append(county_id_key[0])

            if state and state not in states:
                states.append(state)
            if county_id_key and state:
                county_id_key_states_dict[county_id_key] = state
            county_votes_dict[county_id_key] = votes
            county_BBvotes_dict[county_id_key] = BB_votes
            county_RRvotes_dict[county_id_key] = RR_votes


##########################################################
# Function/Method Name: read_state_data
#
# Author: Boris Pashkov
#
# Parameters: file_path (str) - The path to the state-level CSV file.
#
# Return Value: state_data (list) - A list of dictionaries containing
# state-level information, including electoral votes.
#
# Description:
# Reads and processes state-level data from a CSV file, including
# the electoral votes allocated to each state.
##########################################################
def read_state_data(file_path):
    with open(file_path, 'r') as file:
        reader = csv.reader(file)
        for _ in range(4):
            next(reader)

        for row in reader:
            if row[-3]:
                elec_votes = row[-3]
                electoral_votes.append(elec_votes)
            state = row[0]
            state_id = row[1]

            num_of_reg_voters = row[6]

            state_stateid[state] = state_id

            state_num_reg_voters[state] = str((num_of_reg_voters.replace(',', '')).strip())

        for state, elec in zip(electoral_votes, states):
            state_electoral_votes.update({elec: state})

##########################################################
# Function/Method Name: calculate_state_votes
#
# Author: Boris Pashkov
#
# Parameters: county_data (list) - A list of county-level vote data.
#
# Return Value: state_totals (dict) - A dictionary with the total votes
# for each candidate in each state.
#
# Description:
# Calculates the total votes for each candidate (BB and RR) in each state
# by aggregating the data from individual counties.
##########################################################
def calculate_state_votes():
    for state in states:
        states_BB_votes[state] = 0
        states_RR_votes[state] = 0
        states_votes[state] = 0

    for county_id_key in county_id_key_states_dict:
        states_votes[county_id_key_states_dict[county_id_key]] += int(county_votes_dict[county_id_key])
        if county_id_key and county_id_key in county_BBvotes_dict:
            states_BB_votes[county_id_key_states_dict[county_id_key]] += int(county_BBvotes_dict[county_id_key])

        if county_id_key and county_id_key in county_RRvotes_dict:
            states_RR_votes[county_id_key_states_dict[county_id_key]] += int(county_RRvotes_dict[county_id_key])

##########################################################
# Function/Method Name: calculate_winner_percentage
#
# Author: Boris Pashkov
#
# Parameters: state_totals (dict) - A dictionary containing the total votes
# for each candidate in each state.
#
# Return Value: state_percentages (dict) - A dictionary containing the
# percentage of votes for each candidate in each state.
#
# Description:
# Computes the percentage of total votes that each candidate received in
# each state.
##########################################################
def calculate_winner_percentage():
    for state in states:
        if states_BB_votes[state] > states_RR_votes[state]:
            win_perc = round(((states_BB_votes[state] / states_votes[state]) * 100), 2)
            canidate_winner = 'Buggs Bunny'
            state_canidate_winner_key = (state, canidate_winner)
            states_votes_win_perc[state_canidate_winner_key] = win_perc
        elif states_RR_votes[state] > states_BB_votes[state]:
            win_perc = round(((states_RR_votes[state] / states_votes[state]) * 100), 2)
            canidate_winner = 'Road Runner'
            state_canidate_winner_key = (state, canidate_winner)
            states_votes_win_perc[state_canidate_winner_key] = win_perc

##########################################################
# Function/Method Name: determine_popular_vote_winner
#
# Author: Boris Pashkov
#
# Parameters: None
#
# Return Value: winner (str) - The name of the candidate who won the
# popular vote (either "BB" or "RR").
#
# Description:
# Determines the candidate with the highest total vote count across
# all states, thus identifying the popular vote winner.
##########################################################
def calculate_popular_vote():
    BB_tot_votes = 0
    for state in states_BB_votes:
        BB_tot_votes += int(states_BB_votes[state])

    RR_tot_votes = 0
    for state in states_RR_votes:
        RR_tot_votes += int(states_RR_votes[state])

    if BB_tot_votes > RR_tot_votes:
        pop_vote_winner = "Buggs Bunny"
        vote_total = BB_tot_votes
    else:
        pop_vote_winner = "Road Runner"
        vote_total = RR_tot_votes
    print(f'{pop_vote_winner} has won the Popular vote with {vote_total} votes! ')

##########################################################
# Function/Method Name: calculate_electoral_votes
#
# Author: Boris Pashkov
#
# Parameters: state_totals (dict), state_data (list) - The total votes
# in each state and state-level electoral votes data.
#
# Return Value: None
#
# Description:
# Awards electoral votes to the winner of each state based on the total
# votes, storing the electoral votes for each candidate. As well as determining
#who won the electoral college
##########################################################
def calculate_electoral_votes():
    # Task 5: Award Electoral Votes for each state to the winner of the state with the highest number of votes BB or RR.
    BB_electoral_votes = []
    RR_electoral_votes = []

    for state_canidate_winner_key in states_votes_win_perc:
        if state_canidate_winner_key[1] == 'Buggs Bunny':
            BB_electoral_votes.append(int(state_electoral_votes[state_canidate_winner_key[0]]))
        elif state_canidate_winner_key[1] == 'Road Runner':
            RR_electoral_votes.append((int(state_electoral_votes[state_canidate_winner_key[0]])))

    # Task 6: Determine who won the Electoral Vote tally

    if sum(BB_electoral_votes) > sum(RR_electoral_votes):
        electoral_vote_winner = "Buggs Bunny"
        electoral_votes_tally = sum(BB_electoral_votes)
    elif sum(RR_electoral_votes) > sum(BB_electoral_votes):
        electoral_vote_winner = "Road Runner"
        electoral_votes_tally = sum(RR_electoral_votes)

    print(f'{electoral_vote_winner} has won the electoral college with {electoral_votes_tally} electoral votes!')

##########################################################
# Function/Method Name: create_and_print_summary
#
# Author: Michael Cummings
#
# Parameters: None
#
# Return Value: None
#
# Description:
# Generates and prints a summary of the election results, including the
# total votes for each candidate, the percentage of votes, and electoral
# votes awarded per state.
##########################################################
def create_and_print_sumary():
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "Election Summary By State Info"

    ws3.merge_cells('A1:I1')
    ws3['A1'] = 'Summary of Election Results By State'
    ws3.append(['State Name', 'state ID', 'Total Votes', 'Sum of BB Votes', 'Sum of RR Votes', 'Number of Counties',
                'Percent of State Votes', ' ', 'Electroal Votes Awarded', ' '])
    ws3.merge_cells('G2:H2')
    ws3.merge_cells('I2:J2')
    ws3['G3'] = 'BuggsBunny'
    ws3['H3'] = 'RoadRunner'
    ws3['I3'] = 'BuggsBunny'
    ws3['J3'] = 'RoadRunner'

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        # Merge cells from row 2 to row 3 in the current column
        ws3.merge_cells(f'{col}2:{col}3')  # Format: 'ColumnRowStart:ColumnRowEnd'

    for state in states:
        state_num_counties[state] = 0

    for county in county_state_dict:
        state = county_state_dict[county]
        if state in state_num_counties:
            state_num_counties[state] += 1



    for state in states:
        # canidate_win_perc_key = (BB, RR)
        canidate_win_perc_key = (round(((states_BB_votes[state] / states_votes[state]) * 100), 2), round(((states_RR_votes[state] / states_votes[state]) * 100), 2))
        state_canidate_win_perc_key[state] = canidate_win_perc_key



    for state_canidate_winner_key in states_votes_win_perc:
        if state_canidate_winner_key[1] == 'Buggs Bunny':
            state_canidate_elec_wins[state_canidate_winner_key] = int(
                state_electoral_votes[state_canidate_winner_key[0]])
            state_canidate_elec_wins[(state_canidate_winner_key[0], 'Road Runner')] = 0
        elif state_canidate_winner_key[1] == 'Road Runner':
            state_canidate_elec_wins[state_canidate_winner_key] = int(
                state_electoral_votes[state_canidate_winner_key[0]])
            state_canidate_elec_wins[(state_canidate_winner_key[0], 'Buggs Bunny')] = 0

    for state in states:
        total_votes = states_votes[state]
        sum_of_BB_votes = states_BB_votes[state]
        sum_of_RR_votes = states_RR_votes[state]
        state_ID = state_stateid[state]
        num_of_counties = state_num_counties[state]
        Bb_win_perc = state_canidate_win_perc_key[state][0]
        Rr_win_perc = state_canidate_win_perc_key[state][1]

        Bb_elec_votes = state_canidate_elec_wins[(state, "Buggs Bunny")]
        Rr_elec_votes = state_canidate_elec_wins[(state, 'Road Runner')]

        # Task 8: Print summary chart
        ws3.append([state, state_ID, total_votes, sum_of_BB_votes, sum_of_RR_votes, num_of_counties, f'{Bb_win_perc}%', f'{Rr_win_perc}%', Bb_elec_votes, Rr_elec_votes])

        # Calculate totals for the columns
        total_total_votes = sum(states_votes[state] for state in states)
        total_BB_votes = sum(states_BB_votes[state] for state in states)
        total_RR_votes = sum(states_RR_votes[state] for state in states)
        total_num_counties = sum(state_num_counties[state] for state in states)
        total_BB_elec_votes = sum(state_canidate_elec_wins.get((state, 'Buggs Bunny'), 0) for state in states)
        total_RR_elec_votes = sum(state_canidate_elec_wins.get((state, 'Road Runner'), 0) for state in states)
        total_BB_votes_percent = round(((total_BB_votes/total_total_votes) * 100), 2)
        total_RR_votes_percent = round(((total_RR_votes/total_total_votes) * 100), 2)

        # Append the totals row at the bottom
    ws3.append(['Grand Total', '', total_total_votes, total_BB_votes, total_RR_votes, total_num_counties, f'{total_BB_votes_percent}%', f'{total_RR_votes_percent}%', total_BB_elec_votes, total_RR_elec_votes])

    # Save the workbook
    wb3.save("/Users/michaelcummings/Library/Mobile Documents/com~apple~CloudDocs/state_election_results_v.8.xlsx")

##########################################################
# Function/Method Name: create_and_print_summary_for_specified_date
#
# Author: Michael Cummings
#
# Parameters: date_input (str) - The user-provided date and time string.
#
# Return Value: filtered_counties (list) - A list of counties whose data
# was reported before the provided date/time.
#
# Description:
# Filters the counties' results based on the user-specified date/time
# and returns only the counties that have completed their reporting.
##########################################################
def create_and_print_summary_for_specified_date():
    # Loop over each county and reformat the date
    for county in county_return_date:
        if county:
            # Extract the original date string, assuming it is in MM/DD/YYYY format
            date_str = county_return_date[county]
            date_str = date_str.strip()

            # Split the date string into components (MM, DD, YYYY)
            month, day, year = date_str.split('/')

            # Reformat the date as YYYY-MM-DD

            correct_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"

            # Store the corrected date in the dictionary
            county_return_date_correct_format[county] = correct_date


    for county in county_return_time:
        if county:
            # Extract the original date string, assuming it is in MM/DD/YYYY format
            time_str = county_return_time[county]
            time_str = time_str.strip()

            # Split the date string into components (MM, DD, YYYY)
            hour, min, sec = time_str.split(':')

            # Reformat the date as YYYY-MM-DD

            correct_time = f"{hour.zfill(2)}:{min.zfill(2)}:{sec.zfill(2)}"

            # Store the corrected date in the dictionary
            county_return_time_correct_format[county] = correct_time


    result_report_cutoff_input = (input("Input a time for summary results in YYYY-MM-DD HH:MM:SS format")).strip()

    try:
        result_report_cutoff = datetime.strptime(result_report_cutoff_input, "%Y-%m-%d %H:%M:%S")
        print(f"Result report cutoff: {result_report_cutoff}")
    except ValueError:
        print("Could not load summary for inputted datetime. Please input datetime in correct format")
        exit()  # Exit the program if the input format is incorrect

    for county in county_return_date_correct_format and county_return_time_correct_format:

        datetime_str = str(county_return_date_correct_format[county]) + ' ' + str(
            county_return_time_correct_format[county])  #
        county_datetime = datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")

        # Compare the date of the county to the cutoff
        if county_datetime < result_report_cutoff:
            valid_counties.append(county)


    for state in states:
        state_Buggs_votes[state] = 0
        state_Runner_votes[state] = 0

    for county in valid_counties:
        if county in county_BBvotes_dict:
            state_Buggs_votes[county_state_dict[county]] += int(county_BBvotes_dict[county])
        if county in county_RRvotes_dict:
            state_Runner_votes[county_state_dict[county]] += int(county_RRvotes_dict[county])

    Buggs_total_votes = 0
    for state in state_Buggs_votes:
        Buggs_total_votes += state_Buggs_votes[state]

    Runner_total_votes = 0
    for state in state_Runner_votes:
        Runner_total_votes += state_Runner_votes[state]

    if Buggs_total_votes > Runner_total_votes:
        B = "Buggs bunny was leading in the popular"
    else:
        B = "Road Runner was leading in the popular vote"


    for state in states:
        state_list_county[state] = []

    for county in county_state_dict:
        if county_state_dict[county] in state_list_county:
            state_list_county[county_state_dict[county]].append(county)

    for state in states:
        if len(state_list_county[state]) == state_num_counties[state]:
            valid_electoral_states.append(state)


    for state in valid_electoral_states:
        if states_BB_votes[state] > states_RR_votes[state]:
            BB_electoral_votes_v2.append(int(state_electoral_votes[state]))
        elif states_RR_votes[state] > states_BB_votes[state]:
            RR_electoral_votes_v2.append(int(state_electoral_votes[state]))

    if sum(BB_electoral_votes_v2) > sum(RR_electoral_votes_v2):
        A = "Buggs bunny was leading in the electoral college"
    else:
        A = "Road Runner was leading in the electoral college"

    print(f"At {result_report_cutoff}. {B}.")
    print(f"At {result_report_cutoff}. {A}.")


    for state in states:
        state_valid_county_votes[state] = 0

    for county in valid_counties:
        if county_state_dict[county] in state_valid_county_votes:
            state_valid_county_votes[county_state_dict[county]] += int(county_votes_not_key[county])


    for state in states:
        state_perc_votes_placed[state] = f"{round((int(state_valid_county_votes[state]) / int(state_num_reg_voters[state]) * 100), 2)}% of votes have been placed"

    print(state_perc_votes_placed)

##########################################################
# Function/Method Name: plot_county_results_for_country
#
# Author: Michael Cummings
#
# Parameters: None
#
# Return Value: None
#
# Description:
# Visualizes the county-level election results across the entire country
# using matplotlib, with color coding for Buggs Bunny (green) and Road
# Runner (magenta) as the winners in each county.
##########################################################
def plot_county_results_for_country():
    fig, ax = plt.subplots(figsize=(10, 10))

    for county in county_lat_long_key:
        lat = float(county_lat_long_key[county][0])
        lon = float(county_lat_long_key[county][1])

        if county_BBvotes_not_key[county] > county_RRvotes_not_key[county]:
            color = 'green'
        else:
            color = 'magenta'

        ax.scatter(lon, lat, color=color, s=5)

    ax.set_title("U.S. County Results", fontsize=16)
    ax.set_xlabel("Longitude")
    ax.set_ylabel("Latitude")

    plt.show()

##########################################################
# Function/Method Name: plot_county_results_for_specified_state
#
# Author: Michael Cummings
#
# Parameters: state_id (str) - The ID of the state to visualize results for.
#
# Return Value: None
#
# Description:
# Visualizes the county-level election results for a specified state,
# showing which candidate (BB or RR) won in each county.
##########################################################
def plot_county_results_for_specified_state():
    state_id_chosen = input("Select a state to display it's counties:").upper()

    state_name = None

    for state, state_id in state_stateid.items():
        if state_id_chosen == state_id:
            state_name = state
    counties_to_display = state_list_county[state_name]

    fig, bx = plt.subplots(figsize=(10, 10))

    for county in counties_to_display:
        latitude = float(county_lat_long_key[county][0])
        longitude = float(county_lat_long_key[county][1])

        if county_BBvotes_not_key[county] > county_RRvotes_not_key[county]:
            color = 'green'
        else:
            color = 'magenta'

        bx.scatter(longitude, latitude, color=color, s=5)

    bx.set_title(f"{state_name} County Results", fontsize=16)
    bx.set_xlabel("Longitude")
    bx.set_ylabel("Latitude")

    plt.show()


def main():
    # Main script execution flow
    read_county_data('/Users/michaelcummings/Library/Mobile Documents/com~apple~CloudDocs/Voting-Counties.csv')

    state_electoral_votes = read_state_data('/Users/michaelcummings/Library/Mobile Documents/com~apple~CloudDocs/State-Info.csv')

    calculate_state_votes()
    calculate_winner_percentage()


    calculate_popular_vote()
    calculate_electoral_votes()

    create_and_print_sumary()
    create_and_print_summary_for_specified_date()

    plot_county_results_for_country()
    plot_county_results_for_specified_state()


if __name__ == '__main__':
    main()
